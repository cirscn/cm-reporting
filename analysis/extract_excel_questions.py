#!/usr/bin/env python3
"""
Extract question text + DV options from Excel Declaration sheet.
Outputs analysis/excel_question_options.json
"""
from __future__ import annotations

import json
import re
from pathlib import Path
from typing import Dict, List, Tuple

import openpyxl
from openpyxl.utils.cell import range_boundaries

ROOT = Path(__file__).resolve().parents[1]
TEMPLATE_ROOT = ROOT / "app" / "templates"
OUTPUT = ROOT / "analysis" / "excel_question_options.json"

QUESTION_RE = re.compile(r"^(\d+)\)")
COMPANY_RE = re.compile(r"^([A-I])[\).]")
ANSWER_HEADERS = {"回答", "Answer"}


def iter_templates() -> List[Tuple[str, Path]]:
    items = []
    for folder in sorted(TEMPLATE_ROOT.iterdir()):
        if folder.is_dir():
            items.append((folder.name, folder))
    return items


def dv_list_values(wb: openpyxl.Workbook, sheet: str, formula1: str) -> List[str]:
    if not formula1:
        return []
    formula = str(formula1).strip()
    if "!" in formula:
        sheet_part, range_part = formula.split("!", 1)
        sheet_name = sheet_part.strip("'")
    else:
        sheet_name = sheet
        range_part = formula
    range_part = range_part.replace("$", "")
    if ":" not in range_part:
        # literal list like "Yes,No" is uncommon in these templates
        if "," in range_part:
            return [v.strip() for v in range_part.split(",") if v.strip()]
        return []
    try:
        min_col, min_row, max_col, max_row = range_boundaries(range_part)
    except ValueError:
        return []
    if sheet_name not in wb.sheetnames:
        return []
    ws = wb[sheet_name]
    values: List[str] = []
    for row in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
        for cell in row:
            val = cell.value
            if val is None:
                continue
            text = str(val).strip()
            if text:
                values.append(text)
    return values


def build_dv_index(ws) -> List[openpyxl.worksheet.datavalidation.DataValidation]:
    return [dv for dv in ws.data_validations.dataValidation if dv.type == "list"]


def find_dv_for_cell(dvs, cell) -> List[str]:
    for dv in dvs:
        for sqref in str(dv.sqref).split():
            sqref = sqref.replace("$", "")
            if ":" in sqref:
                min_col, min_row, max_col, max_row = range_boundaries(sqref)
                if min_row <= cell.row <= max_row and min_col <= cell.column <= max_col:
                    return [dv.formula1]
            else:
                if sqref == cell.coordinate:
                    return [dv.formula1]
    return []


def first_text_cell(ws, row: int, max_col: int = 10) -> Tuple[int, str]:
    for col in range(1, max_col + 1):
        val = ws.cell(row, col).value
        if isinstance(val, str) and val.strip():
            return col, val.strip()
    return 0, ""


def extract_questions(ws, wb):
    dvs = build_dv_index(ws)
    result = {"questions": [], "company": []}
    question_rows: List[int] = []
    company_rows: List[int] = []

    for r in range(1, 200):
        for c in range(1, 6):
            val = ws.cell(r, c).value
            if isinstance(val, str):
                text = val.strip()
                if QUESTION_RE.match(text):
                    question_rows.append(r)
                if COMPANY_RE.match(text):
                    company_rows.append(r)
        # early stop if both sections found and passed
    question_rows = sorted(set(question_rows))
    company_rows = sorted(set(company_rows))

    # Map question rows
    for idx, r in enumerate(question_rows):
        q_col, q_text = first_text_cell(ws, r)
        m = QUESTION_RE.match(q_text)
        if not m:
            continue
        q_num = m.group(1)
        # find answer column in header row
        answer_col = None
        for c in range(1, 12):
            if ws.cell(r, c).value in ANSWER_HEADERS:
                answer_col = c
                break
        if answer_col is None:
            answer_col = 4
        # pick first data row below
        data_row = r + 1
        # skip blank rows
        while data_row < r + 20:
            _, v = first_text_cell(ws, data_row)
            if v:
                break
            data_row += 1
        cell = ws.cell(data_row, answer_col)
        formulas = find_dv_for_cell(dvs, cell)
        dv_values = []
        for f in formulas:
            dv_values = dv_list_values(wb, ws.title, f)
            if dv_values:
                break
        result["questions"].append(
            {
                "number": q_num,
                "text": q_text,
                "answer_col": answer_col,
                "dv": dv_values,
            }
        )

    # Map company questions
    for r in company_rows:
        q_col, q_text = first_text_cell(ws, r)
        m = COMPANY_RE.match(q_text)
        if not m:
            continue
        key = m.group(1)
        answer_col = None
        for c in range(1, 12):
            if ws.cell(r, c).value in ANSWER_HEADERS:
                answer_col = c
                break
        if answer_col is None:
            answer_col = 4
        cell = ws.cell(r, answer_col)
        formulas = find_dv_for_cell(dvs, cell)
        dv_values = []
        for f in formulas:
            dv_values = dv_list_values(wb, ws.title, f)
            if dv_values:
                break
        result["company"].append(
            {
                "key": key,
                "text": q_text,
                "answer_col": answer_col,
                "dv": dv_values,
            }
        )

    return result


def main() -> None:
    output: Dict[str, Dict[str, dict]] = {}
    for template, folder in iter_templates():
        for path in sorted(folder.glob("*.xlsx")):
            if path.name.startswith("~$"):
                continue
            version_match = re.search(r"_(\d+(?:\.\d+)*)\.xlsx$", path.name)
            if not version_match:
                continue
            version = version_match.group(1)
            wb = openpyxl.load_workbook(path, data_only=True)
            if "Declaration" not in wb.sheetnames:
                wb.close()
                continue
            ws = wb["Declaration"]
            data = extract_questions(ws, wb)
            output.setdefault(template, {})[version] = data
            wb.close()

    OUTPUT.write_text(json.dumps(output, ensure_ascii=False, indent=2), encoding="utf-8")


if __name__ == "__main__":
    main()
