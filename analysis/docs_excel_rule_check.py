#!/usr/bin/env python3
"""
Validate docs vs Excel for key rules across Declaration / Smelter List / Mine List / Product List / Checker
/ Minerals Scope / Instructions / Definitions.

Outputs:
- docs/diffs/docs-excel-rule-check.md (pass/fail summary + evidence)
- docs/diffs/docs-excel-evidence.md (evidence index: Excel -> docs line)
"""
from __future__ import annotations

import argparse
import csv
import json
import re
from dataclasses import dataclass
from pathlib import Path
from typing import Callable, Iterable, List, Optional, Tuple

import openpyxl
from openpyxl.utils.cell import range_boundaries

ROOT = Path(__file__).resolve().parents[1]
TEMPLATE_ROOT = ROOT / "app" / "templates"
DOCS_ROOT = ROOT / "docs"
ANALYSIS_ROOT = ROOT / "analysis"
SCAN_PREFIX = "scan-"
VERSION_RE = re.compile(r"_(\d+(?:\.\d+)*)\.xlsx$")


@dataclass
class Rule:
    rule_id: str
    description: str
    excel_evidence: Callable[[], Tuple[bool, str]]
    doc_patterns: List[re.Pattern]
    doc_paths: List[Path]
    templates: Tuple[str, ...] = ()
    versions: Tuple[str, ...] = ()


def find_latest_scan(root: Path) -> Path:
    scans = sorted([p for p in root.iterdir() if p.is_dir() and p.name.startswith(SCAN_PREFIX)])
    if not scans:
        raise FileNotFoundError("No scan-YYYY-MM-DD directories found in analysis/.")
    return scans[-1]


def parse_version_tuple(v: str) -> Tuple[int, ...]:
    return tuple(int(x) for x in v.split("."))


def list_template_versions(template: str) -> List[str]:
    folder = TEMPLATE_ROOT / template.upper()
    versions = []
    for xlsx in folder.glob("*.xlsx"):
        m = VERSION_RE.search(xlsx.name)
        if not m:
            continue
        versions.append(m.group(1))
    return sorted(set(versions), key=parse_version_tuple)


def load_template_path(template: str, version: str) -> Path:
    folder = TEMPLATE_ROOT / template.upper()
    for xlsx in folder.glob("*.xlsx"):
        m = VERSION_RE.search(xlsx.name)
        if not m:
            continue
        if m.group(1) == version:
            return xlsx
    raise FileNotFoundError(f"Template {template} {version} not found")


def norm_text(val) -> str:
    if val is None:
        return ""
    if isinstance(val, float) and val.is_integer():
        val = int(val)
    text = str(val)
    return text.replace("\r\n", "\n").replace("\r", "\n").strip()


def load_dv_catalog(scan_root: Path, template: str) -> List[dict]:
    path = scan_root / f"{template.lower()}_rule_catalog_dv.csv"
    if not path.exists():
        return []
    rows: List[dict] = []
    with path.open(encoding="utf-8") as f:
        reader = csv.DictReader(f)
        for row in reader:
            rows.append(row)
    return rows


def cell_in_sqref(cell: str, sqref: str) -> bool:
    target_col = re.sub(r"\d+", "", cell)
    target_row = int(re.sub(r"\D+", "", cell))
    for part in sqref.split():
        part = part.replace("$", "")
        if not part:
            continue
        if ":" in part:
            min_col, min_row, max_col, max_row = range_boundaries(part)
            # convert letter to index via range_boundaries already
            if min_row <= target_row <= max_row:
                # compare column letters by index
                if min_col <= openpyxl.utils.cell.column_index_from_string(target_col) <= max_col:
                    return True
        else:
            if part == cell or part == cell.replace("$", ""):
                return True
    return False


def dv_list_for_cell(dv_rows: List[dict], version: str, sheet: str, cell: str) -> List[dict]:
    hits = []
    for row in dv_rows:
        if row.get("version") != version:
            continue
        if row.get("sheet") != sheet:
            continue
        if row.get("type") != "list":
            continue
        sqref = row.get("sqref", "")
        if cell_in_sqref(cell, sqref):
            hits.append(row)
    return hits


def dv_list_formulas_for_cell(dv_rows: List[dict], version: str, sheet: str, cell: str) -> List[str]:
    hits = dv_list_for_cell(dv_rows, version, sheet, cell)
    formulas = []
    for row in hits:
        formula1 = row.get("formula1", "")
        if formula1:
            formulas.append(formula1)
    return formulas


def dv_list_values(wb: openpyxl.Workbook, sheet: str, formula1: str) -> List[str]:
    if not formula1:
        return []
    formula = formula1.strip()
    if "!" in formula:
        sheet_part, range_part = formula.split("!", 1)
        sheet_part = sheet_part.strip("'")
        sheet_name = sheet_part
    else:
        sheet_name = sheet
        range_part = formula
    range_part = range_part.replace("$", "")
    if ":" not in range_part:
        return []
    try:
        min_col, min_row, max_col, max_row = range_boundaries(range_part)
    except ValueError:
        return []
    if sheet_name not in wb.sheetnames:
        return []
    ws = wb[sheet_name]
    values = []
    for row in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
        for cell in row:
            values.append(norm_text(cell.value))
    return [v for v in values if v]


def l_sheet_value(template: str, version: str, sheet: str, cell: str) -> str:
    wb = openpyxl.load_workbook(load_template_path(template, version), data_only=False)
    if "L" not in wb.sheetnames:
        return ""
    ws = wb["L"]
    for row in ws.iter_rows(min_row=1, values_only=True):
        if len(row) < 5:
            continue
        if row[1] == sheet and row[2] == cell:
            return norm_text(row[3])
    return ""


def sheet_presence(template: str, version: str, required: List[str]) -> Tuple[bool, str]:
    wb = openpyxl.load_workbook(load_template_path(template, version), data_only=False)
    missing = [name for name in required if name not in wb.sheetnames]
    ok = not missing
    evid = f"{template} {version} missing sheets={missing}" if missing else f"{template} {version} sheets ok"
    return ok, evid


def sheet_presence_all_versions(template: str, expected_fn: Callable[[str], List[str]]) -> Tuple[bool, str]:
    versions = list_template_versions(template)
    if not versions:
        return False, f"{template} no versions found"
    missing_map = []
    for version in versions:
        wb = openpyxl.load_workbook(load_template_path(template, version), data_only=False)
        expected = expected_fn(version)
        missing = [name for name in expected if name not in wb.sheetnames]
        if missing:
            missing_map.append(f"{version}: {missing}")
    ok = not missing_map
    evid = f"{template} all versions ok" if ok else f"{template} missing={'; '.join(missing_map)}"
    return ok, evid


def sheet_absence(template: str, version: str, disallowed: List[str]) -> Tuple[bool, str]:
    wb = openpyxl.load_workbook(load_template_path(template, version), data_only=False)
    present = [name for name in disallowed if name in wb.sheetnames]
    ok = not present
    evid = f"{template} {version} present disallowed={present}" if present else f"{template} {version} disallowed ok"
    return ok, evid


def instructions_english_only(template: str, version: str) -> Tuple[bool, str]:
    wb = openpyxl.load_workbook(load_template_path(template, version), data_only=False)
    if "L" not in wb.sheetnames:
        return False, f"{template} {version} L sheet missing"
    ws = wb["L"]
    found = False
    sample = ""
    for row in ws.iter_rows(min_row=1, values_only=True):
        if len(row) < 5:
            continue
        if row[1] != "Instructions":
            continue
        en = norm_text(row[3])
        if re.search(r"ENGLISH only", en, re.IGNORECASE):
            found = True
            sample = en[:160]
            break
    evid = f"{template} {version} ENGLISH only found={found}; sample={sample!r}"
    return found, evid


def instructions_minerals_scope_recommendation(template: str, version: str) -> Tuple[bool, str]:
    wb = openpyxl.load_workbook(load_template_path(template, version), data_only=False)
    if "L" not in wb.sheetnames:
        return False, f"{template} {version} L sheet missing"
    ws = wb["L"]
    found = ""
    for row in ws.iter_rows(min_row=1, values_only=True):
        if len(row) < 5:
            continue
        if row[1] != "Instructions":
            continue
        en = norm_text(row[3])
        if "Minerals Scope" in en and re.search(r"strongly recommended", en, re.IGNORECASE):
            found = en
            break
    ok = bool(found)
    evid = f"{template} {version} Minerals Scope recommendation found={ok}; sample={found[:160]!r}"
    return ok, evid


def definitions_contains_cahra(template: str, version: str) -> Tuple[bool, str]:
    wb = openpyxl.load_workbook(load_template_path(template, version), data_only=True)
    if "Definitions" not in wb.sheetnames:
        return False, f"{template} {version} Definitions sheet missing"
    ws = wb["Definitions"]
    found = None
    for row in ws.iter_rows(min_row=1, values_only=True):
        for val in row:
            text = norm_text(val)
            if not text:
                continue
            if "CAHRA" in text or "受冲突影响和高风险地区" in text:
                found = text
                break
        if found:
            break
    ok = bool(found)
    sample = (found[:160] + "...") if found and len(found) > 160 else (found or "")
    evid = f"{template} {version} Definitions CAHRA found={ok}; sample={sample!r}"
    return ok, evid


def instructions_date_format(template: str, version: str) -> Tuple[bool, str]:
    wb = openpyxl.load_workbook(load_template_path(template, version), data_only=False)
    if "L" not in wb.sheetnames:
        return False, f"{template} {version} L sheet missing"
    ws = wb["L"]
    found = ""
    for row in ws.iter_rows(min_row=1, values_only=True):
        if len(row) < 5:
            continue
        if row[1] != "Instructions":
            continue
        en = norm_text(row[3])
        if "Date of Completion" in en and "DD-MMM-YYYY" in en:
            found = en
            break
    ok = bool(found)
    evid = f"{template} {version} date format found={ok}; sample={found[:160]!r}"
    return ok, evid


def instructions_filename_example(template: str, version: str) -> Tuple[bool, str]:
    wb = openpyxl.load_workbook(load_template_path(template, version), data_only=False)
    if "L" not in wb.sheetnames:
        return False, f"{template} {version} L sheet missing"
    ws = wb["L"]
    found = ""
    for row in ws.iter_rows(min_row=1, values_only=True):
        if len(row) < 5:
            continue
        if row[1] != "Instructions":
            continue
        en = norm_text(row[3])
        if "companyname-date" in en and "YYYY-MM-DD" in en:
            found = en
            break
    ok = bool(found)
    evid = f"{template} {version} filename example found={ok}; sample={found[:160]!r}"
    return ok, evid


def definitions_contains_terms(template: str, version: str, terms: List[str]) -> Tuple[bool, str]:
    wb = openpyxl.load_workbook(load_template_path(template, version), data_only=True)
    if "Definitions" not in wb.sheetnames:
        return False, f"{template} {version} Definitions sheet missing"
    ws = wb["Definitions"]
    found = {t: False for t in terms}
    for row in ws.iter_rows(min_row=1, values_only=True):
        for val in row:
            text = norm_text(val)
            if not text:
                continue
            for t in terms:
                if t in text:
                    found[t] = True
    missing = [t for t, ok in found.items() if not ok]
    ok = not missing
    evid = f"{template} {version} missing_terms={missing}" if missing else f"{template} {version} terms ok"
    return ok, evid


def definitions_terms_multi(template: str, versions: List[str], terms: List[str]) -> Tuple[bool, str]:
    details = []
    ok = True
    for version in versions:
        hit, evid = definitions_contains_terms(template, version, terms)
        details.append(evid)
        if not hit:
            ok = False
    return ok, " / ".join(details)


def instructions_date_format_multi(template: str, versions: List[str]) -> Tuple[bool, str]:
    details = []
    ok = True
    for version in versions:
        hit, evid = instructions_date_format(template, version)
        details.append(evid)
        if not hit:
            ok = False
    return ok, " / ".join(details)


def instructions_filename_example_multi(template: str, versions: List[str]) -> Tuple[bool, str]:
    details = []
    ok = True
    for version in versions:
        hit, evid = instructions_filename_example(template, version)
        details.append(evid)
        if not hit:
            ok = False
    return ok, " / ".join(details)

def load_cf_formulas(scan_root: Path) -> List[dict]:
    path = scan_root / "xml-cf-report.jsonl"
    if not path.exists():
        return []
    rows = []
    with path.open(encoding="utf-8") as f:
        for line in f:
            if not line.strip():
                continue
            rows.append(json.loads(line))
    return rows


def find_doc_evidence(patterns: List[re.Pattern], paths: Iterable[Path]) -> Optional[str]:
    evidences: List[str] = []
    seen = set()
    for pat in patterns:
        found = None
        for path in paths:
            if not path.exists():
                continue
            lines = path.read_text(encoding="utf-8").splitlines()
            for idx, line in enumerate(lines, 1):
                if pat.search(line):
                    found = f"{path}:{idx}: {line.strip()}"
                    break
            if found:
                break
        if found:
            if found not in seen:
                evidences.append(found)
                seen.add(found)
        else:
            evidences.append(f"NOT FOUND: {pat.pattern}")
    if not evidences:
        return None
    return "<br>".join(evidences)


def rule_matches(rule: Rule, templates: Optional[set], versions: Optional[set]) -> bool:
    if templates:
        if rule.templates and not (set(rule.templates) & templates):
            return False
        if not rule.templates:
            return False
    if versions:
        if rule.versions and not (set(rule.versions) & versions):
            return False
        if not rule.versions:
            return False
    return True


def build_rules(scan_root: Path) -> List[Rule]:
    dv_emrt = load_dv_catalog(scan_root, "EMRT")
    dv_amrt = load_dv_catalog(scan_root, "AMRT")
    dv_cmrt = load_dv_catalog(scan_root, "CMRT")
    cf_rows = load_cf_formulas(scan_root)

    def emrt_mine_sheet() -> Tuple[bool, str]:
        wb20 = openpyxl.load_workbook(load_template_path("EMRT", "2.0"), data_only=False)
        has_mine = "Mine List" in wb20.sheetnames
        return has_mine, f"EMRT 2.0 sheets include Mine List={has_mine}"

    def emrt_mine_dropdown_21() -> Tuple[bool, str]:
        hits_21 = dv_list_for_cell(dv_emrt, "2.1", "Mine List", "B5")
        hits_20 = dv_list_for_cell(dv_emrt, "2.0", "Mine List", "B5")
        ok = bool(hits_21) and not hits_20
        evid = f"EMRT 2.1 Mine List B5 list dv={bool(hits_21)}; EMRT 2.0 Mine List B5 list dv={bool(hits_20)}"
        return ok, evid

    def amrt_mine_dropdown_13() -> Tuple[bool, str]:
        hits_13 = dv_list_for_cell(dv_amrt, "1.3", "Mine List", "B5")
        hits_12 = dv_list_for_cell(dv_amrt, "1.2", "Mine List", "B5")
        hits_11 = dv_list_for_cell(dv_amrt, "1.1", "Mine List", "B5")
        ok = bool(hits_13) and not hits_12 and not hits_11
        evid = (
            f"AMRT 1.3 Mine List B5 list dv={bool(hits_13)}; "
            f"1.2={bool(hits_12)}; 1.1={bool(hits_11)}"
        )
        return ok, evid

    def mine_bang_cf_emrt21() -> Tuple[bool, str]:
        formulas = []
        for row in cf_rows:
            if row.get("template") != "EMRT" or row.get("version") != "2.1":
                continue
            if row.get("sheet") != "Mine List":
                continue
            for formula in row.get("formulas", []) or []:
                if "!" in formula and ("FIND(\"!\"" in formula or "SEARCH(\"!\"" in formula):
                    formulas.append(formula)
        ok = bool(formulas)
        evid = " ; ".join(formulas[:2]) if formulas else "no CF formula with FIND/SEARCH '!'"
        return ok, evid

    def emrt_1x_drc_mica_options() -> Tuple[bool, str]:
        wb = openpyxl.load_workbook(load_template_path("EMRT", "1.3"), data_only=False)
        ws = wb["L"]
        found_drc = False
        found_mica = False
        for row in ws.iter_rows(min_row=1, values_only=True):
            if len(row) < 5:
                continue
            en = norm_text(row[3])
            if "DRC only" in en:
                found_drc = True
            if "India and/or Madagascar only" in en:
                found_mica = True
        ok = found_drc and found_mica
        evid = f"EMRT 1.3 L sheet contains DRC only={found_drc}; India and/or Madagascar only={found_mica}"
        return ok, evid

    def emrt_did_not_survey() -> Tuple[bool, str]:
        wb20 = openpyxl.load_workbook(load_template_path("EMRT", "2.0"), data_only=False)
        wb13 = openpyxl.load_workbook(load_template_path("EMRT", "1.3"), data_only=False)
        # collect list values for Declaration DV lists
        def list_values(wb, version: str) -> List[str]:
            values: List[str] = []
            for row in dv_emrt:
                if row.get("version") != version:
                    continue
                if row.get("sheet") != "Declaration":
                    continue
                if row.get("type") != "list":
                    continue
                formula1 = row.get("formula1", "")
                vals = dv_list_values(wb, "Declaration", formula1)
                values.extend(vals)
            return list(dict.fromkeys(values))

        v20 = list_values(wb20, "2.0")
        v13 = list_values(wb13, "1.3")
        has_20 = any(v == "Did not survey" for v in v20)
        has_13 = any(v == "Did not survey" for v in v13)
        ok = has_20 and not has_13
        evid = f"EMRT 2.0 has Did not survey={has_20}; EMRT 1.3 has Did not survey={has_13}"
        return ok, evid

    def emrt_100_percent_numeric() -> Tuple[bool, str]:
        wb = openpyxl.load_workbook(load_template_path("EMRT", "1.3"), data_only=False)
        ws = wb["L"]
        value = None
        for row in ws.iter_rows(min_row=1, values_only=True):
            if len(row) < 5:
                continue
            if row[1] == "Declaration" and row[2] == "B89":
                value = row[3]
                break
        ok = value == 1
        evid = f"EMRT 1.3 L sheet Declaration B89 English value={value!r}"
        return ok, evid

    def emrt_decl_p38_p39_labels() -> Tuple[bool, str]:
        b38 = l_sheet_value("EMRT", "1.3", "Declaration", "B38")
        b39 = l_sheet_value("EMRT", "1.3", "Declaration", "B39")
        wb = openpyxl.load_workbook(load_template_path("EMRT", "1.3"), data_only=False)
        ws = wb["Declaration"]
        p38 = norm_text(ws["P38"].value)
        p39 = norm_text(ws["P39"].value)
        ok = ("Cobalt" in b38) and ("Mica" in b39) and bool(p38) and bool(p39)
        evid = f"EMRT 1.3 Declaration B38={b38!r}, B39={b39!r}; P38={p38!r}; P39={p39!r}"
        return ok, evid

    def amrt_minerals_scope_intro() -> Tuple[bool, str]:
        b5 = l_sheet_value("AMRT", "1.3", "Minerals Scope", "B5")
        ok = ("Minerals Scope tab" in b5) and ("AMRT requester" in b5)
        evid = f"AMRT 1.3 Minerals Scope B5={b5!r}"
        return ok, evid

    def amrt_minerals_scope_headers() -> Tuple[bool, str]:
        wb = openpyxl.load_workbook(load_template_path("AMRT", "1.3"), data_only=True)
        if "Minerals Scope" not in wb.sheetnames:
            return False, "AMRT 1.3 Minerals Scope sheet missing"
        ws = wb["Minerals Scope"]
        b7 = norm_text(ws["B7"].value)
        c7 = norm_text(ws["C7"].value)
        ok = ("Select Minerals/Metals in Scope" in b7) and ("Reasons for inclusion on the AMRT" in c7)
        evid = f"AMRT 1.3 Minerals Scope B7={b7!r}; C7={c7!r}"
        return ok, evid

    def amrt_minerals_scope_headers_all_versions() -> Tuple[bool, str]:
        expected = {
            "1.1": "Reasons for inclusion on the PRT",
            "1.2": "Reasons for inclusion on the AMRT",
            "1.3": "Reasons for inclusion on the AMRT",
        }
        details = []
        ok = True
        for version, expected_c7 in expected.items():
            wb = openpyxl.load_workbook(load_template_path("AMRT", version), data_only=True)
            if "Minerals Scope" not in wb.sheetnames:
                details.append(f"AMRT {version} Minerals Scope sheet missing")
                ok = False
                continue
            ws = wb["Minerals Scope"]
            b7 = norm_text(ws["B7"].value)
            c7 = norm_text(ws["C7"].value)
            hit = ("Select Minerals/Metals in Scope" in b7) and (expected_c7 in c7)
            details.append(f"AMRT {version} B7={b7!r}; C7={c7!r}")
            if not hit:
                ok = False
        return ok, " / ".join(details)

    def cmrt_smelter_lookup_dropdown() -> Tuple[bool, str]:
        formulas = dv_list_formulas_for_cell(dv_cmrt, "6.5", "Smelter List", "C5")
        ok = any(formula == "SN" for formula in formulas)
        evid = f"CMRT 6.5 Smelter List C5 list formulas={formulas or '[]'}"
        return ok, evid

    def emrt_smelter_lookup_dropdown() -> Tuple[bool, str]:
        formulas = dv_list_formulas_for_cell(dv_emrt, "2.1", "Smelter List", "C5")
        ok = any(formula == "SN" for formula in formulas)
        evid = f"EMRT 2.1 Smelter List C5 list formulas={formulas or '[]'}"
        return ok, evid

    def amrt_smelter_lookup_13() -> Tuple[bool, str]:
        f13 = dv_list_formulas_for_cell(dv_amrt, "1.3", "Smelter List", "C5")
        f12 = dv_list_formulas_for_cell(dv_amrt, "1.2", "Smelter List", "C5")
        ok = any(formula == "SN" for formula in f13) and not any(formula == "SN" for formula in f12)
        evid = f"AMRT 1.3 Smelter List C5 formulas={f13 or '[]'}; 1.2 formulas={f12 or '[]'}"
        return ok, evid

    def emrt_product_requester_fields() -> Tuple[bool, str]:
        d5_21 = l_sheet_value("EMRT", "2.1", "Product List", "D5")
        d5_20 = l_sheet_value("EMRT", "2.0", "Product List", "D5")
        ok = ("Requester Product Number" in d5_21) and ("Comments" in d5_20)
        evid = f"EMRT 2.1 Product List D5={d5_21!r}; EMRT 2.0 Product List D5={d5_20!r}"
        return ok, evid

    def amrt_product_requester_fields() -> Tuple[bool, str]:
        d5_13 = l_sheet_value("AMRT", "1.3", "Product List", "D5")
        e5_13 = l_sheet_value("AMRT", "1.3", "Product List", "E5")
        d5_12 = l_sheet_value("AMRT", "1.2", "Product List", "D5")
        ok = ("Requester Product Number" in d5_13) and ("Requester Product Name" in e5_13) and ("Comments" in d5_12)
        evid = f"AMRT 1.3 Product List D5={d5_13!r}, E5={e5_13!r}; AMRT 1.2 Product List D5={d5_12!r}"
        return ok, evid

    def checker_required_flags_latest() -> Tuple[bool, str]:
        targets = [
            ("CMRT", "6.5"),
            ("EMRT", "2.1"),
            ("CRT", "2.21"),
            ("AMRT", "1.3"),
        ]
        details = []
        ok = True
        for template, version in targets:
            wb = openpyxl.load_workbook(load_template_path(template, version), data_only=True)
            if "Checker" not in wb.sheetnames:
                details.append(f"{template} {version} Checker missing")
                ok = False
                continue
            ws = wb["Checker"]
            found = False
            for row in ws.iter_rows(min_row=1, max_col=6, values_only=True):
                if row[5] == 1:
                    found = True
                    break
            details.append(f"{template} {version} Checker F=1 found={found}")
            if not found:
                ok = False
        return ok, "; ".join(details)

    def emrt_definitions_terms_2x() -> Tuple[bool, str]:
        return definitions_terms_multi("EMRT", ["2.0", "2.1"], ["授权人", "尽职调查"])

    def emrt_definitions_terms_1x() -> Tuple[bool, str]:
        return definitions_terms_multi(
            "EMRT",
            ["1.1", "1.11", "1.2", "1.3"],
            [
                "授权人",
                "尽职调查",
                "独立的私营审核机构",
                "有意添加",
                "有目的添加",
                "冲突矿产",
                "受冲突影响和高风险地区",
                "经济合作与发展组织",
                "冶炼厂",
            ],
        )

    def amrt_definitions_terms_all() -> Tuple[bool, str]:
        return definitions_terms_multi("AMRT", ["1.1", "1.2", "1.3"], ["授权人", "尽职调查"])

    def cmrt_definitions_terms_all() -> Tuple[bool, str]:
        return definitions_terms_multi(
            "CMRT",
            ["6.01", "6.1", "6.22", "6.31", "6.4", "6.5"],
            ["授权人", "有意添加", "独立的第三方审核机构", "冲突矿产", "受冲突影响和高风险地区", "经济合作与发展组织", "冶炼厂"],
        )

    def crt_definitions_terms_all() -> Tuple[bool, str]:
        return definitions_terms_multi(
            "CRT",
            ["2.2", "2.21"],
            ["授权人", "尽职调查", "独立的私营审核机构", "有意添加", "冲突矿产", "受冲突影响和高风险地区", "经济合作与发展组织", "冶炼厂"],
        )

    def emrt_instructions_date_format_2x() -> Tuple[bool, str]:
        return instructions_date_format_multi("EMRT", ["2.0", "2.1"])

    def emrt_instructions_filename_2x() -> Tuple[bool, str]:
        return instructions_filename_example_multi("EMRT", ["2.0", "2.1"])

    return [
        Rule(
            rule_id="CMRT-SHEETS-LATEST",
            description="CMRT 最新版本包含所有核心 sheet",
            excel_evidence=lambda: sheet_presence(
                "CMRT",
                "6.5",
                ["Revision", "Instructions", "Definitions", "Declaration", "Smelter List", "Checker", "Product List", "Smelter Look-up"],
            ),
            doc_patterns=[re.compile(r"CMRT：Revision / Instructions / Definitions / Declaration / Smelter List / Checker / Product List / Smelter Look-up")],
            doc_paths=[DOCS_ROOT / "prd" / "conflict-minerals-prd.md", DOCS_ROOT / "diffs" / "cross-template.md"],
            templates=("CMRT",),
            versions=("6.5",),
        ),
        Rule(
            rule_id="EMRT-SHEETS-LATEST",
            description="EMRT 最新版本包含 Mine List 且包含 Smelter Look-up",
            excel_evidence=lambda: sheet_presence(
                "EMRT",
                "2.1",
                ["Revision", "Instructions", "Definitions", "Declaration", "Smelter List", "Checker", "Product List", "Smelter Look-up", "Mine List"],
            ),
            doc_patterns=[re.compile(r"EMRT：Revision / Instructions / Definitions / Declaration / Smelter List / Checker / Product List / Smelter Look-up（2\.0\+ 追加 Mine List）")],
            doc_paths=[DOCS_ROOT / "prd" / "conflict-minerals-prd.md", DOCS_ROOT / "diffs" / "cross-template.md"],
            templates=("EMRT",),
            versions=("2.1",),
        ),
        Rule(
            rule_id="CRT-SHEETS-LATEST",
            description="CRT 最新版本包含所有核心 sheet",
            excel_evidence=lambda: sheet_presence(
                "CRT",
                "2.21",
                ["Revision", "Instructions", "Definitions", "Declaration", "Smelter List", "Checker", "Product List", "Smelter Look-up"],
            ),
            doc_patterns=[re.compile(r"CRT：Instructions / Revision / Definitions / Declaration / Smelter List / Checker / Product List / Smelter Look-up")],
            doc_paths=[DOCS_ROOT / "prd" / "conflict-minerals-prd.md", DOCS_ROOT / "diffs" / "cross-template.md"],
            templates=("CRT",),
            versions=("2.21",),
        ),
        Rule(
            rule_id="AMRT-SHEETS-1.3",
            description="AMRT 1.3 包含 Minerals Scope / Mine List / Smelter Look-up",
            excel_evidence=lambda: sheet_presence(
                "AMRT",
                "1.3",
                ["Revision", "Instructions", "Definitions", "Declaration", "Minerals Scope", "Smelter List", "Checker", "Mine List", "Product List", "Smelter Look-up"],
            ),
            doc_patterns=[re.compile(r"AMRT：Revision / Instructions / Definitions / Declaration / Minerals Scope / Smelter List / Checker / Mine List / Product List / Smelter Look-up")],
            doc_paths=[DOCS_ROOT / "prd" / "conflict-minerals-prd.md", DOCS_ROOT / "diffs" / "cross-template.md"],
            templates=("AMRT",),
            versions=("1.3",),
        ),
        Rule(
            rule_id="AMRT-SMELTER-LOOKUP-ABSENCE-1.2",
            description="AMRT 1.2 不包含 Smelter Look-up",
            excel_evidence=lambda: (
                sheet_presence(
                    "AMRT",
                    "1.2",
                    ["Revision", "Instructions", "Definitions", "Declaration", "Minerals Scope", "Smelter List", "Checker", "Mine List", "Product List"],
                )[0]
                and sheet_absence("AMRT", "1.2", ["Smelter Look-up"])[0],
                "; ".join(
                    [
                        sheet_presence(
                            "AMRT",
                            "1.2",
                            ["Revision", "Instructions", "Definitions", "Declaration", "Minerals Scope", "Smelter List", "Checker", "Mine List", "Product List"],
                        )[1],
                        sheet_absence("AMRT", "1.2", ["Smelter Look-up"])[1],
                    ]
                ),
            ),
            doc_patterns=[re.compile(r"1\.2 无 Smelter Look-up|1\.1/1\.2：无 Smelter Look-up")],
            doc_paths=[DOCS_ROOT / "diffs" / "amrt.md", DOCS_ROOT / "diffs" / "acceptance" / "amrt-1.2.md"],
            templates=("AMRT",),
            versions=("1.2",),
        ),
        Rule(
            rule_id="INSTRUCTIONS-ENGLISH-ONLY",
            description="Instructions 明示“ENGLISH only”",
            excel_evidence=lambda: (
                instructions_english_only("CMRT", "6.5")[0]
                and instructions_english_only("EMRT", "2.1")[0]
                and instructions_english_only("CRT", "2.21")[0]
                and instructions_english_only("AMRT", "1.3")[0],
                " / ".join(
                    [
                        instructions_english_only("CMRT", "6.5")[1],
                        instructions_english_only("EMRT", "2.1")[1],
                        instructions_english_only("CRT", "2.21")[1],
                        instructions_english_only("AMRT", "1.3")[1],
                    ]
                ),
            ),
            doc_patterns=[re.compile(r"仅以英文作答")],
            doc_paths=[DOCS_ROOT / "prd" / "conflict-minerals-prd.md"],
            templates=("CMRT", "EMRT", "CRT", "AMRT"),
            versions=("6.5", "2.1", "2.21", "1.3"),
        ),
        Rule(
            rule_id="AMRT-MINERALS-SCOPE-INTRO",
            description="Minerals Scope 页签说明（AMRT 申请人可选填）",
            excel_evidence=amrt_minerals_scope_intro,
            doc_patterns=[re.compile(r"AMRT 申请人可选择填写该页签以提供更多细节")],
            doc_paths=[DOCS_ROOT / "diffs" / "amrt.md"],
            templates=("AMRT",),
            versions=("1.3",),
        ),
        Rule(
            rule_id="AMRT-MINERALS-SCOPE-HEADERS",
            description="Minerals Scope 表头（Select Minerals/Metals in Scope / Reasons for inclusion on the AMRT）",
            excel_evidence=amrt_minerals_scope_headers,
            doc_patterns=[re.compile(r"Select Minerals/Metals in Scope"), re.compile(r"Reasons for inclusion on the AMRT")],
            doc_paths=[DOCS_ROOT / "diffs" / "amrt.md", DOCS_ROOT / "diffs" / "cross-template.md"],
            templates=("AMRT",),
            versions=("1.3",),
        ),
        Rule(
            rule_id="AMRT-MINERALS-SCOPE-HEADERS-ALL",
            description="AMRT 1.1/1.2/1.3 Minerals Scope 表头差异（PRT vs AMRT）",
            excel_evidence=amrt_minerals_scope_headers_all_versions,
            doc_patterns=[re.compile(r"Select Minerals/Metals in Scope"), re.compile(r"PRT 字样")],
            doc_paths=[DOCS_ROOT / "diffs" / "amrt.md", DOCS_ROOT / "diffs" / "cross-template.md"],
            templates=("AMRT",),
            versions=("1.1", "1.2", "1.3"),
        ),
        Rule(
            rule_id="AMRT-MINERALS-SCOPE-RECOMMEND",
            description="Instructions 建议先完成 Minerals Scope 再提交",
            excel_evidence=lambda: instructions_minerals_scope_recommendation("AMRT", "1.3"),
            doc_patterns=[re.compile(r"强烈建议先填写 Minerals Scope")],
            doc_paths=[DOCS_ROOT / "diffs" / "amrt.md"],
            templates=("AMRT",),
            versions=("1.3",),
        ),
        Rule(
            rule_id="EMRT-INSTRUCTIONS-DATE-FORMAT",
            description="Instructions 日期格式为 DD-MMM-YYYY",
            excel_evidence=lambda: instructions_date_format("EMRT", "1.3"),
            doc_patterns=[re.compile(r"DD-MMM-YYYY")],
            doc_paths=[DOCS_ROOT / "diffs" / "emrt.md", DOCS_ROOT / "prd" / "conflict-minerals-prd.md"],
            templates=("EMRT",),
            versions=("1.3",),
        ),
        Rule(
            rule_id="EMRT-INSTRUCTIONS-DATE-FORMAT-2X",
            description="EMRT 2.0/2.1 Instructions 日期格式为 DD-MMM-YYYY",
            excel_evidence=emrt_instructions_date_format_2x,
            doc_patterns=[re.compile(r"DD-MMM-YYYY")],
            doc_paths=[DOCS_ROOT / "diffs" / "emrt.md", DOCS_ROOT / "prd" / "conflict-minerals-prd.md"],
            templates=("EMRT",),
            versions=("2.0", "2.1"),
        ),
        Rule(
            rule_id="EMRT-INSTRUCTIONS-FILENAME-EXAMPLE",
            description="Instructions 文件名示例 companyname-date.xlsx",
            excel_evidence=lambda: instructions_filename_example("EMRT", "1.3"),
            doc_patterns=[re.compile(r"companyname-date\.xlsx"), re.compile(r"YYYY-MM-DD")],
            doc_paths=[DOCS_ROOT / "diffs" / "emrt.md"],
            templates=("EMRT",),
            versions=("1.3",),
        ),
        Rule(
            rule_id="EMRT-INSTRUCTIONS-FILENAME-EXAMPLE-2X",
            description="EMRT 2.0/2.1 Instructions 文件名示例 companyname-date.xlsx",
            excel_evidence=emrt_instructions_filename_2x,
            doc_patterns=[re.compile(r"companyname-date\.xlsx"), re.compile(r"YYYY-MM-DD")],
            doc_paths=[DOCS_ROOT / "diffs" / "emrt.md"],
            templates=("EMRT",),
            versions=("2.0", "2.1"),
        ),
        Rule(
            rule_id="DEFINITIONS-CAHRA",
            description="Definitions 中包含 CAHRA 定义",
            excel_evidence=lambda: definitions_contains_cahra("EMRT", "1.3"),
            doc_patterns=[re.compile(r"CAHRA")],
            doc_paths=[
                DOCS_ROOT / "prd" / "appendix-definitions.md",
                DOCS_ROOT / "prd" / "definitions-master.md",
            ],
            templates=("EMRT",),
            versions=("1.3",),
        ),
        Rule(
            rule_id="EMRT-DEFINITIONS-TERMS-2X",
            description="EMRT 2.0/2.1 Definitions 含“授权人/尽职调查”",
            excel_evidence=emrt_definitions_terms_2x,
            doc_patterns=[re.compile(r"授权人"), re.compile(r"尽职调查")],
            doc_paths=[DOCS_ROOT / "prd" / "appendix-definitions.md", DOCS_ROOT / "prd" / "definitions-master.md"],
            templates=("EMRT",),
            versions=("2.0", "2.1"),
        ),
        Rule(
            rule_id="EMRT-DEFINITIONS-TERMS-1X",
            description="EMRT 1.x Definitions 核心术语覆盖（授权人/尽职调查/独立私营审核/有意添加/有目的添加/冲突矿产/CAHRA/OECD/冶炼厂）",
            excel_evidence=emrt_definitions_terms_1x,
            doc_patterns=[
                re.compile(r"授权人"),
                re.compile(r"尽职调查"),
                re.compile(r"独立的私营审核机构"),
                re.compile(r"有意添加"),
                re.compile(r"有目的添加"),
                re.compile(r"冲突矿产"),
                re.compile(r"受冲突影响和高风险地区"),
                re.compile(r"经济合作与发展组织"),
                re.compile(r"冶炼厂"),
            ],
            doc_paths=[DOCS_ROOT / "prd" / "appendix-definitions.md", DOCS_ROOT / "prd" / "definitions-master.md"],
            templates=("EMRT",),
            versions=("1.1", "1.11", "1.2", "1.3"),
        ),
        Rule(
            rule_id="AMRT-DEFINITIONS-TERMS-ALL",
            description="AMRT 1.1/1.2/1.3 Definitions 含“授权人/尽职调查”",
            excel_evidence=amrt_definitions_terms_all,
            doc_patterns=[re.compile(r"授权人"), re.compile(r"尽职调查")],
            doc_paths=[DOCS_ROOT / "prd" / "appendix-definitions.md", DOCS_ROOT / "prd" / "definitions-master.md"],
            templates=("AMRT",),
            versions=("1.1", "1.2", "1.3"),
        ),
        Rule(
            rule_id="CMRT-DEFINITIONS-TERMS-ALL",
            description="CMRT 6.x Definitions 核心术语覆盖（授权人/有意添加/独立审核/冲突矿产/CAHRA/OECD/冶炼厂）",
            excel_evidence=cmrt_definitions_terms_all,
            doc_patterns=[
                re.compile(r"授权人"),
                re.compile(r"有意添加"),
                re.compile(r"独立的第三方审核机构"),
                re.compile(r"冲突矿产"),
                re.compile(r"受冲突影响和高风险地区"),
                re.compile(r"经济合作与发展组织"),
                re.compile(r"冶炼厂"),
            ],
            doc_paths=[DOCS_ROOT / "prd" / "appendix-definitions.md", DOCS_ROOT / "prd" / "definitions-master.md"],
            templates=("CMRT",),
            versions=("6.01", "6.1", "6.22", "6.31", "6.4", "6.5"),
        ),
        Rule(
            rule_id="CRT-DEFINITIONS-TERMS-ALL",
            description="CRT 2.2/2.21 Definitions 核心术语覆盖（授权人/尽职调查/独立私营审核/有意添加/冲突矿产/CAHRA/OECD/冶炼厂）",
            excel_evidence=crt_definitions_terms_all,
            doc_patterns=[
                re.compile(r"授权人"),
                re.compile(r"尽职调查"),
                re.compile(r"独立的私营审核机构"),
                re.compile(r"有意添加"),
                re.compile(r"冲突矿产"),
                re.compile(r"受冲突影响和高风险地区"),
                re.compile(r"经济合作与发展组织"),
                re.compile(r"冶炼厂"),
            ],
            doc_paths=[DOCS_ROOT / "prd" / "appendix-definitions.md", DOCS_ROOT / "prd" / "definitions-master.md"],
            templates=("CRT",),
            versions=("2.2", "2.21"),
        ),
        Rule(
            rule_id="EMRT-MINE-SHEET-2X",
            description="EMRT 2.0+ 存在 Mine List sheet",
            excel_evidence=emrt_mine_sheet,
            doc_patterns=[re.compile(r"Mine List（矿厂清单）"), re.compile(r"2\.0 起")],
            doc_paths=[DOCS_ROOT / "diffs" / "emrt.md"],
            templates=("EMRT",),
            versions=("2.0", "2.1"),
        ),
        Rule(
            rule_id="EMRT-MINE-DROPDOWN-2.1",
            description="EMRT 2.1 Mine List B列为下拉；2.0 为手填",
            excel_evidence=emrt_mine_dropdown_21,
            doc_patterns=[re.compile(r"2\.1.*下拉"), re.compile(r"2\.0.*手动")],
            doc_paths=[DOCS_ROOT / "diffs" / "emrt.md"],
            templates=("EMRT",),
            versions=("2.0", "2.1"),
        ),
        Rule(
            rule_id="AMRT-MINE-DROPDOWN-1.3",
            description="AMRT 1.3 Mine List B列为下拉；1.1/1.2 为手填",
            excel_evidence=amrt_mine_dropdown_13,
            doc_patterns=[re.compile(r"Mine List.*下拉"), re.compile(r"1\.1/1\.2.*手动|1\.1/1\.2.*手填")],
            doc_paths=[DOCS_ROOT / "diffs" / "amrt.md"],
            templates=("AMRT",),
            versions=("1.1", "1.2", "1.3"),
        ),
        Rule(
            rule_id="MINE-BANG-CF",
            description="Mine List 中包含 \"!\" 触发提示（条件格式）",
            excel_evidence=mine_bang_cf_emrt21,
            doc_patterns=[re.compile(r"包含.*!.*触发提示")],
            doc_paths=[DOCS_ROOT / "diffs" / "emrt.md", DOCS_ROOT / "diffs" / "amrt.md"],
            templates=("EMRT", "AMRT"),
            versions=("2.1", "1.3", "1.2", "1.1"),
        ),
        Rule(
            rule_id="EMRT-1X-DRC-MICA",
            description="EMRT 1.x Q3 选项：Cobalt=DRC only；Mica=India and/or Madagascar only",
            excel_evidence=emrt_1x_drc_mica_options,
            doc_patterns=[re.compile(r"DRC only"), re.compile(r"India and/or Madagascar only")],
            doc_paths=[DOCS_ROOT / "diffs" / "emrt.md"],
            templates=("EMRT",),
            versions=("1.1", "1.11", "1.2", "1.3"),
        ),
        Rule(
            rule_id="EMRT-DID-NOT-SURVEY",
            description="EMRT 2.0+ Q2 含 Did not survey；1.x 不含",
            excel_evidence=emrt_did_not_survey,
            doc_patterns=[re.compile(r"Did not survey")],
            doc_paths=[DOCS_ROOT / "diffs" / "emrt.md"],
            templates=("EMRT",),
            versions=("2.0", "2.1", "1.3"),
        ),
        Rule(
            rule_id="EMRT-1X-100PCT-NUMERIC",
            description="EMRT 1.x Q2 的 100% 在模板中为数值 1",
            excel_evidence=emrt_100_percent_numeric,
            doc_patterns=[re.compile(r"100%.*对应值为 1")],
            doc_paths=[DOCS_ROOT / "diffs" / "emrt.md"],
            templates=("EMRT",),
            versions=("1.1", "1.11", "1.2", "1.3"),
        ),
        Rule(
            rule_id="EMRT-DECL-P38-P39-COBALT-MICA",
            description="EMRT 1.3 Declaration P38/P39 对应 Cobalt/Mica 行标记",
            excel_evidence=emrt_decl_p38_p39_labels,
            doc_patterns=[re.compile(r"Q3 是否从 CAHRA 采购指定矿产"), re.compile(r"Cobalt 回答"), re.compile(r"Mica 回答")],
            doc_paths=[DOCS_ROOT / "diffs" / "emrt.md"],
            templates=("EMRT",),
            versions=("1.3",),
        ),
        Rule(
            rule_id="CMRT-SMELTER-LOOKUP",
            description="CMRT Smelter List 冶炼厂查找为下拉（SN）",
            excel_evidence=cmrt_smelter_lookup_dropdown,
            doc_patterns=[re.compile(r"冶炼厂查找为下拉字段")],
            doc_paths=[DOCS_ROOT / "diffs" / "cmrt.md"],
            templates=("CMRT",),
            versions=("6.5",),
        ),
        Rule(
            rule_id="EMRT-SMELTER-LOOKUP",
            description="EMRT Smelter List 冶炼厂查找为下拉（SN）",
            excel_evidence=emrt_smelter_lookup_dropdown,
            doc_patterns=[re.compile(r"冶炼厂查找为下拉字段")],
            doc_paths=[DOCS_ROOT / "diffs" / "emrt.md"],
            templates=("EMRT",),
            versions=("2.1",),
        ),
        Rule(
            rule_id="AMRT-SMELTER-LOOKUP-1.3",
            description="AMRT 1.3 Smelter List 冶炼厂名称（下拉）启用，1.1/1.2 无",
            excel_evidence=amrt_smelter_lookup_13,
            doc_patterns=[re.compile(r"1\.3：新增 Smelter Look-up"), re.compile(r"1\.1/1\.2：无 Smelter Look-up")],
            doc_paths=[DOCS_ROOT / "diffs" / "amrt.md"],
            templates=("AMRT",),
            versions=("1.1", "1.2", "1.3"),
        ),
        Rule(
            rule_id="EMRT-PRODUCT-REQUESTER",
            description="EMRT 2.1 Product List 新增请求方字段；2.0 无",
            excel_evidence=emrt_product_requester_fields,
            doc_patterns=[re.compile(r"请求方的产品编号"), re.compile(r"Product List 表头变化")],
            doc_paths=[DOCS_ROOT / "diffs" / "emrt.md"],
            templates=("EMRT",),
            versions=("2.0", "2.1"),
        ),
        Rule(
            rule_id="AMRT-PRODUCT-REQUESTER",
            description="AMRT 1.3 Product List 新增请求方字段；1.2 无",
            excel_evidence=amrt_product_requester_fields,
            doc_patterns=[re.compile(r"请求方的产品编号"), re.compile(r"Product List 表头变化")],
            doc_paths=[DOCS_ROOT / "diffs" / "amrt.md"],
            templates=("AMRT",),
            versions=("1.2", "1.3"),
        ),
        Rule(
            rule_id="CHECKER-REQUIRED-FLAGS",
            description="Checker 明确必填项（F=1）存在",
            excel_evidence=checker_required_flags_latest,
            doc_patterns=[re.compile(r"Checker 明确必填项"), re.compile(r"F=1")],
            doc_paths=[DOCS_ROOT / "diffs" / "checker-matrix.md"],
            templates=("CMRT", "EMRT", "CRT", "AMRT"),
            versions=("6.5", "2.1", "2.21", "1.3"),
        ),
    ]


def main() -> None:
    parser = argparse.ArgumentParser(description="Validate docs vs Excel key rules.")
    parser.add_argument("--scan-dir", type=str, default="", help="Use a specific scan directory (analysis/scan-YYYY-MM-DD).")
    parser.add_argument("--templates", type=str, default="", help="Comma-separated templates to include (CMRT,EMRT,CRT,AMRT).")
    parser.add_argument("--versions", type=str, default="", help="Comma-separated versions to include (e.g., 2.1,2.0,1.3).")
    args = parser.parse_args()

    scan_root = Path(args.scan_dir) if args.scan_dir else find_latest_scan(ANALYSIS_ROOT)
    rules = build_rules(scan_root)

    templates_filter = {t.strip().upper() for t in args.templates.split(",") if t.strip()} or None
    versions_filter = {v.strip() for v in args.versions.split(",") if v.strip()} or None

    if templates_filter or versions_filter:
        rules = [rule for rule in rules if rule_matches(rule, templates_filter, versions_filter)]

    # Build reports
    report_path = DOCS_ROOT / "diffs" / "docs-excel-rule-check.md"
    evidence_path = DOCS_ROOT / "diffs" / "docs-excel-evidence.md"

    report_lines = []
    evidence_lines = []

    report_lines.append(
        "# Docs ↔ Excel 规则校验（Declaration / Smelter List / Mine List / Product List / Checker / Minerals Scope / Instructions / Definitions）"
    )
    report_lines.append("")
    report_lines.append(f"> 扫描源：{scan_root}")
    if templates_filter:
        report_lines.append(f"> 模板过滤：{', '.join(sorted(templates_filter))}")
    if versions_filter:
        report_lines.append(f"> 版本过滤：{', '.join(sorted(versions_filter))}")
    report_lines.append("")
    report_lines.append("| Rule | 结果 | Excel 证据 | Docs 证据 |")
    report_lines.append("|---|---|---|---|")

    evidence_lines.append("# 审计证据清单（Excel → Docs）")
    evidence_lines.append("")
    evidence_lines.append(f"> 扫描源：{scan_root}")
    if templates_filter:
        evidence_lines.append(f"> 模板过滤：{', '.join(sorted(templates_filter))}")
    if versions_filter:
        evidence_lines.append(f"> 版本过滤：{', '.join(sorted(versions_filter))}")
    evidence_lines.append("")
    evidence_lines.append("| Rule | Excel 证据 | Docs 证据 |")
    evidence_lines.append("|---|---|---|")

    for rule in rules:
        ok, excel_evid = rule.excel_evidence()
        doc_evid = find_doc_evidence(rule.doc_patterns, rule.doc_paths)
        doc_evid_text = doc_evid or "NOT FOUND"
        report_lines.append(f"| {rule.rule_id} | {'PASS' if ok else 'FAIL'} | {excel_evid} | {doc_evid_text} |")
        evidence_lines.append(f"| {rule.rule_id} | {excel_evid} | {doc_evid_text} |")

    report_path.write_text("\n".join(report_lines) + "\n", encoding="utf-8")
    evidence_path.write_text("\n".join(evidence_lines) + "\n", encoding="utf-8")
    print(f"written {report_path}")
    print(f"written {evidence_path}")


if __name__ == "__main__":
    main()
