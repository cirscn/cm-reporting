#!/usr/bin/env python3
"""
Manual evidence review for non-Product List sheets.

Outputs:
- manual-sheet-review-diff.md (per-version evidence)
- manual-sheet-review-fixlist.md (mismatches + summary matrix)
"""

from __future__ import annotations

import argparse
import re
from dataclasses import dataclass
from pathlib import Path
from typing import Dict, Iterable, List, Optional, Tuple

import openpyxl

ROOT = Path(__file__).resolve().parents[1]
TEMPLATE_ROOT = ROOT / "app" / "templates"
VERSION_RE = re.compile(r"_(\d+(?:\.\d+)*)$")


@dataclass(frozen=True)
class Signal:
    key: str
    sheet_patterns: List[re.Pattern]
    doc_patterns: List[re.Pattern]
    note: str


def parse_version(name: str) -> Tuple[str, Tuple[int, ...]]:
    match = VERSION_RE.search(name)
    if not match:
        return "", (0,)
    version = match.group(1)
    parts = [int(part) for part in version.split(".") if part]
    return version, tuple(parts) or (0,)


def load_template_versions() -> Dict[str, List[Tuple[str, Tuple[int, ...], Path]]]:
    templates: Dict[str, List[Tuple[str, Tuple[int, ...], Path]]] = {}
    for folder in TEMPLATE_ROOT.iterdir():
        if not folder.is_dir():
            continue
        entries = []
        for xlsx in folder.glob("*.xlsx"):
            if xlsx.name.startswith("~$"):
                continue
            version, key = parse_version(xlsx.stem)
            if not version:
                continue
            entries.append((version, key, xlsx))
        entries.sort(key=lambda x: x[1])
        templates[folder.name] = entries
    return templates


def doc_paths_for(template: str, version: str, sheet: str) -> List[Path]:
    lower = template.lower()
    paths: List[Path] = []
    accept = ROOT / "docs" / "diffs" / "acceptance" / f"{lower}-{version}.md"
    if accept.exists():
        paths.append(accept)
    if template == "AMRT" and version == "1.1":
        legacy = ROOT / "docs" / "diffs" / "amrt-1.1.md"
        if legacy.exists():
            paths.append(legacy)

    base = [
        ROOT / "docs" / "diffs" / f"{lower}.md",
        ROOT / "docs" / "prd" / "field-dictionary.md",
        ROOT / "docs" / "prd" / "conflict-minerals-prd.md",
        ROOT / "docs" / "prd" / "pm-onepager.md",
    ]
    if sheet in {"Declaration", "Checker"}:
        base.extend(
            [
                ROOT / "docs" / "diffs" / "cross-template.md",
                ROOT / "docs" / "diffs" / "checker-matrix.md",
            ]
        )
    if sheet in {"Smelter List", "Mine List"}:
        base.append(ROOT / "docs" / "diffs" / "conditional-format-matrix.md")

    for path in base:
        if path.exists():
            paths.append(path)
    return paths


def find_doc_evidence(paths: Iterable[Path], patterns: List[re.Pattern]) -> str:
    for path in paths:
        if not path.exists():
            continue
        lines = path.read_text(encoding="utf-8").splitlines()
        for idx, line in enumerate(lines, 1):
            for pat in patterns:
                if pat.search(line):
                    return f"{path}:{idx}: {line.strip()}"
    return ""


def find_sheet_evidence(ws, patterns: List[re.Pattern], max_rows: int = 200, max_cols: int = 20) -> str:
    for row in range(1, max_rows + 1):
        for col in range(1, max_cols + 1):
            val = ws.cell(row=row, column=col).value
            if val is None:
                continue
            text = str(val).strip()
            if not text:
                continue
            for pat in patterns:
                if pat.search(text):
                    return f"{ws.cell(row=row, column=col).coordinate}: {text}"
    return ""


def build_signals(template: str, version_key: Tuple[int, ...], sheet: str) -> List[Signal]:
    def v_at_least(target: Tuple[int, ...]) -> bool:
        return version_key >= target

    if sheet == "Declaration":
        signals = [
            Signal(
                key="company_name",
                sheet_patterns=[re.compile(r"公司名称")],
                doc_patterns=[re.compile(r"公司名称")],
                note="company name field",
            ),
            Signal(
                key="scope",
                sheet_patterns=[re.compile(r"申报范围")],
                doc_patterns=[re.compile(r"申报范围")],
                note="scope field",
            ),
            Signal(
                key="scope_description",
                sheet_patterns=[re.compile(r"范围描述")],
                doc_patterns=[re.compile(r"范围描述")],
                note="scope description field",
            ),
            Signal(
                key="effective_date",
                sheet_patterns=[re.compile(r"(生效日期|授权日期)")],
                doc_patterns=[re.compile(r"(生效日期|授权日期)")],
                note="effective/authorization date",
            ),
        ]
        if template == "EMRT" and v_at_least((2, 0)):
            signals.append(
                Signal(
                    key="mineral_scope",
                    sheet_patterns=[re.compile(r"矿产申报范围|选择贵公司的矿产申报范围")],
                    doc_patterns=[re.compile(r"矿产申报范围")],
                    note="mineral scope field (EMRT 2.0+)",
                )
            )
        if template == "AMRT":
            signals.append(
                Signal(
                    key="mineral_scope",
                    sheet_patterns=[re.compile(r"矿产申报范围|选择贵公司的矿产申报范围")],
                    doc_patterns=[re.compile(r"矿产申报范围")],
                    note="mineral scope field (AMRT)",
                )
            )
        return signals

    if sheet == "Smelter List":
        signals = [
            Signal(
                key="metal",
                sheet_patterns=[re.compile(r"金属")],
                doc_patterns=[re.compile(r"金属")],
                note="metal column",
            ),
            Signal(
                key="smelter_name",
                sheet_patterns=[re.compile(r"冶炼厂名称")],
                doc_patterns=[re.compile(r"冶炼厂名称")],
                note="smelter name column",
            ),
            Signal(
                key="country",
                sheet_patterns=[re.compile(r"国家|地区")],
                doc_patterns=[re.compile(r"国家|地区")],
                note="country column",
            ),
        ]
        if template in {"CMRT", "EMRT", "CRT"} or (template == "AMRT" and v_at_least((1, 3))):
            signals.append(
                Signal(
                    key="id_input",
                    sheet_patterns=[re.compile(r"(识别|ID).*号|识别号码")],
                    doc_patterns=[re.compile(r"识别号码|识别号|ID")],
                    note="smelter ID input column",
                )
            )
        return signals

    if sheet == "Mine List":
        return [
            Signal(
                key="metal",
                sheet_patterns=[re.compile(r"金属")],
                doc_patterns=[re.compile(r"金属")],
                note="metal column",
            ),
            Signal(
                key="smelter_name",
                sheet_patterns=[re.compile(r"从该矿厂采购的冶炼厂的名称")],
                doc_patterns=[re.compile(r"从该矿厂采购的冶炼厂的名称")],
                note="smelter name from mine",
            ),
            Signal(
                key="mine_name",
                sheet_patterns=[re.compile(r"矿厂|矿场名称")],
                doc_patterns=[re.compile(r"矿厂|矿场")],
                note="mine name column",
            ),
            Signal(
                key="country",
                sheet_patterns=[re.compile(r"国家|地区")],
                doc_patterns=[re.compile(r"国家|地区")],
                note="country column",
            ),
        ]

    if sheet == "Checker":
        return [
            Signal(
                key="checker_exists",
                sheet_patterns=[re.compile(r".")],
                doc_patterns=[re.compile(r"Checker|校验|必填", re.I)],
                note="checker rules section present",
            )
        ]

    return []


def run_review(out_diff: Path, out_fix: Path) -> int:
    templates = load_template_versions()
    mismatches = 0
    matrix_rows: List[Tuple[str, str, str, str]] = []
    diff_lines: List[str] = []
    fix_lines: List[str] = []

    diff_lines.append("# 手工复核（非 Product List）")
    diff_lines.append("")
    diff_lines.append("- 覆盖范围：`docs/diffs/**` + `docs/prd/**` + `docs/diffs/acceptance/**`")
    diff_lines.append("")

    fix_lines.append("# 文档改写建议清单（手工复核矩阵）")
    fix_lines.append("")
    fix_lines.append("- 覆盖范围：`docs/diffs/**` + `docs/prd/**` + `docs/diffs/acceptance/**`")
    fix_lines.append("")

    for template in sorted(templates):
        for version, version_key, path in templates[template]:
            wb = openpyxl.load_workbook(path, data_only=True)
            for sheet in ["Declaration", "Smelter List", "Mine List", "Checker"]:
                if sheet not in wb.sheetnames:
                    continue
                ws = wb[sheet]
                doc_paths = doc_paths_for(template, version, sheet)
                signals = build_signals(template, version_key, sheet)
                status = "OK"

                diff_lines.append(f"## {template} {version} - {sheet}")
                diff_lines.append("")
                diff_lines.append("| Signal | Sheet Evidence | Doc Evidence | Note |")
                diff_lines.append("| --- | --- | --- | --- |")

                for signal in signals:
                    sheet_ev = find_sheet_evidence(ws, signal.sheet_patterns)
                    doc_ev = find_doc_evidence(doc_paths, signal.doc_patterns)
                    diff_lines.append(
                        f"| {signal.key} | {sheet_ev or '-'} | {doc_ev or '-'} | {signal.note} |"
                    )
                    if bool(sheet_ev) != bool(doc_ev):
                        status = "Mismatch"
                        mismatches += 1
                        fix_lines.append(f"## {template} {version} - {sheet}")
                        fix_lines.append("")
                        fix_lines.append(f"- 缺失/不一致：{signal.note}")
                        if sheet_ev:
                            fix_lines.append(f"  - 模板证据：`{sheet_ev}`")
                        if doc_ev:
                            fix_lines.append(f"  - 文档证据：{doc_ev}")
                        fix_lines.append("  - 建议：补充或修正文档口径，保持原术语。")
                        fix_lines.append("")

                diff_lines.append("")
                matrix_rows.append((template, version, sheet, status))
            wb.close()

    fix_lines.append("## 对照矩阵")
    fix_lines.append("")
    fix_lines.append("| Template | Version | Sheet | Status |")
    fix_lines.append("| --- | --- | --- | --- |")
    for template, version, sheet, status in matrix_rows:
        fix_lines.append(f"| {template} | {version} | {sheet} | {status} |")
    fix_lines.append("")

    if mismatches == 0:
        fix_lines.append("## 结论")
        fix_lines.append("")
        fix_lines.append("- 未发现需要修正的文档项。")
        fix_lines.append("")

    out_diff.write_text("\n".join(diff_lines), encoding="utf-8")
    out_fix.write_text("\n".join(fix_lines), encoding="utf-8")
    print(f"Wrote detailed diff to {out_diff}")
    print(f"Wrote fix list to {out_fix}")
    return 0


def main() -> int:
    parser = argparse.ArgumentParser(description="Manual review for non-Product List sheets.")
    parser.add_argument("--out-diff", type=Path, default=None)
    parser.add_argument("--out-fix", type=Path, default=None)
    args = parser.parse_args()

    out_diff = args.out_diff or (ROOT / "analysis" / "scan-2026-01-24" / "manual-sheet-review-diff.md")
    out_fix = args.out_fix or (ROOT / "analysis" / "scan-2026-01-24" / "manual-sheet-review-fixlist.md")
    return run_review(out_diff, out_fix)


if __name__ == "__main__":
    raise SystemExit(main())
